from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Initialize workbook
wb = Workbook()
ws = wb.active
ws.title = "Sheet Analysis"

# Define column headers
headers = [
    'Sr.No', 'Entry', 'Rtrcmnt Entry %', 'DOWN', 'Qnty', 'Capital', 'Total Quantity',
    'Total Invested', 'First_TGT 1.50%', 'EXIT_1st_HALF Quantity', 'Second_TGT',
    'EXIT_2nd_HALF Quantity', 'AVG_on Capital', 'FINAL_TGT 1.50%', 'Final_TGT AWAY %',
    'First_TGT Profit', 'Second_TGT Profit', 'NET_Profit', 'Capital Gain', 'Loss Running',
    'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC'
]
for col, header in enumerate(headers, 1):
    ws[f"{get_column_letter(col)}1"] = header

# User input for U1
u1_value = float(input("Enter U1 value [e.g., 250000]: "))

# Fill static data
for row in range(2, 83):
    ws[f"A{row}"] = row - 1  # Sr.No
ws["B2"] = 100  # Initial Entry
ws["U1"] = u1_value
ws["V1"] = "=IF(U1=100000,-0.642,IF(U1=150000,-0.428,IF(U1=200000,-0.214,IF(U1=250000,0.002,IF(U1=300000,0.216,IF(U1=350000,0.436,IF(U1=400000,0.6511,IF(U1=450000,0.872,IF(U1=500000,1.09,IF(U1=550000,1.311,IF(U1=600000,1.532,IF(U1=650000,1.754,IF(U1=700000,1.977,IF(U1=750000,2.201,IF(U1=800000,2.425,IF(U1=850000,2.65,IF(U1=900000,2.877,IF(U1=950000,3.104,IF(U1=1000000,3.332,0.023)))))))))))))))))))"
ws["F2"] = 20000
ws["F3"] = 1100

# DOWN values starting from D3 (D2 is 0 or empty)
down_values = (
    [-(i * 0.0025) for i in range(1, 21)] +  # Sr.No 2-21 (rows 3-22): -0.25%
    [-0.055 - (i * 0.005) for i in range(0, 20)] +  # Sr.No 22-41 (rows 23-42): -0.50%
    [-0.1575 - (i * 0.0075) for i in range(0, 14)] +  # Sr.No 42-55 (rows 43-56): -0.75%
    [-0.265 - (i * 0.01) for i in range(0, 15)] +  # Sr.No 56-70 (rows 57-71): -1.00%
    [-0.4175 - (i * 0.0125) for i in range(0, 11)]  # Sr.No 71-81 (rows 72-82): -1.25%
)
for row, down in enumerate(down_values[:80], 3):  # Start from row 3 (D3)
    ws[f"D{row}"] = down

# Row 2 (Sr.No 1)
ws["E2"] = "=ROUND(F2/B2,0)"  # Integer
ws["G2"] = "=E2"
ws["H2"] = "=ROUND(B2*E2,2)"  # 2 decimals
ws["M2"] = "=ROUND(B2,2)"     # 2 decimals
ws["N2"] = "=ROUND(M2*1.5%+M2,2)"  # 2 decimals
ws["O2"] = "=(N2-B2)/B2"            # 2 decimals
ws["R2"] = "=ROUND(S2-H2,2)"       # 2 decimals
ws["S2"] = "=ROUND(H2*1.5%+H2,2)"  # 2 decimals
ws["T2"] = "=ROUND((H2*D2+H2)-H2,2)"  # 2 decimals (D2 is 0 or empty)

# Row 3 (Sr.No 2)
ws["B3"] = "=ABS($B$2*D3+$B$2)"  # 2 decimals
ws["E3"] = "=ROUND(F3/B3,0)"              # Integer
ws["G3"] = "=E2+E3"
ws["H3"] = "=ROUND(H2+F3,2)"              # 2 decimals
ws["M3"] = "=ROUND(H3/G3,2)"              # 2 decimals
ws["N3"] = "=ROUND(M3*1.5%+M3,2)"         # 2 decimals
ws["O3"] = "=(N3-B3)/B3"         # 2 decimals
ws["R3"] = "=ROUND(S3-H3,2)"              # 2 decimals
ws["S3"] = "=ROUND(H3*1.5%+H3,2)"         # 2 decimals
ws["T3"] = "=ROUND((H3*D3+H3)-H3,2)"      # 2 decimals

# Row 4 (Sr.No 3)
ws["B4"] = "=ABS($B$2*D4+$B$2)"  # 2 decimals
ws["E4"] = "=ROUND(F4/B4,0)"              # Integer
ws["F4"] = "=F3+ROUND(F3*$V$1,0)"         # Integer
ws["G4"] = "=G3+E4"
ws["H4"] = "=ROUND(H3+F4,2)"              # 2 decimals
ws["M4"] = "=ROUND(H4/G4,2)"              # 2 decimals
ws["N4"] = "=ROUND(M4*1.5%+M4,2)"         # 2 decimals
ws["O4"] = "=(N4-B4)/B4"                    # 2 decimals
ws["R4"] = "=ROUND(S4-H4,2)"              # 2 decimals
ws["S4"] = "=ROUND(H4*1.5%+H4,2)"         # 2 decimals
ws["T4"] = "=ROUND((H4*D4+H4)-H4,2)"      # 2 decimals

# Rows 5-9 (Sr.No 4-8)
for row in range(5, 10):
    ws[f"B{row}"] = f"=ABS($B$2*D{row}+$B$2)"
    ws[f"E{row}"] = f"=ROUND(F{row}/B{row},0)"
    ws[f"F{row}"] = f"=F{row-1}+ROUND(F{row-1}*0.022,0)"
    ws[f"G{row}"] = f"=G{row-1}+E{row}"
    ws[f"H{row}"] = f"=ROUND(H{row-1}+F{row},2)"
    ws[f"M{row}"] = f"=ROUND(H{row}/G{row},2)"
    ws[f"N{row}"] = f"=ROUND(M{row}*1.5%+M{row},2)"
    ws[f"O{row}"] = f"=(N{row}-B{row})/B{row}"
    ws[f"R{row}"] = f"=ROUND(S{row}-H{row},2)"
    ws[f"S{row}"] = f"=ROUND(H{row}*1.5%+H{row},2)"
    ws[f"T{row}"] = f"=ROUND((H{row}*D{row}+H{row})-H{row},2)"

# Rows 10-22 (Sr.No 9-21)
for row in range(10, 23):
    ws[f"B{row}"] = f"=ABS($B$2*D{row}+$B$2)"
    ws[f"E{row}"] = f"=ROUND(F{row}/B{row},0)"
    ws[f"F{row}"] = f"=F{row-1}+ROUND(F{row-1}*0.022,0)"
    ws[f"G{row}"] = f"=G{row-1}+E{row}"
    ws[f"H{row}"] = f"=ROUND(H{row-1}+F{row},2)"
    ws[f"I{row}"] = f"=ROUND(B{row}*1.5%+B{row},2)"
    ws[f"J{row}"] = f"=ROUND(G{row}/2,0)"
    ws[f"M{row}"] = f"=ROUND(H{row}/G{row},2)"
    ws[f"N{row}"] = f"=ROUND(M{row}*1.5%+M{row},2)"
    ws[f"O{row}"] = f"=(N{row}-B{row})/B{row}"
    ws[f"P{row}"] = f"=(I{row}-B{row})*J{row}"
    ws[f"Q{row}"] = f"=ROUND((K{row}-I{row})*L{row},2)"
    ws[f"R{row}"] = f"=ROUND(S{row}-H{row},2)"
    ws[f"S{row}"] = f"=ROUND(H{row}*1.5%+H{row},2)"
    ws[f"T{row}"] = f"=ROUND((H{row}*D{row}+H{row})-H{row},2)"

# Rows 23-82 (Sr.No 22-81)
for row in range(23, 83):
    ws[f"B{row}"] = f"=ABS($B$2*D{row}+$B$2)"
    ws[f"E{row}"] = f"=ROUND(F{row}/B{row},0)"
    ws[f"F{row}"] = f"=F{row-1}+ROUND(F{row-1}*0.022,0)"
    ws[f"G{row}"] = f"=G{row-1}+E{row}"
    ws[f"H{row}"] = f"=ROUND(H{row-1}+F{row},2)"
    ws[f"I{row}"] = f"=ROUND(B{row}*1.5%+B{row},2)"
    ws[f"J{row}"] = f"=ROUND(G{row}/2,0)"
    ws[f"K{row}"] = f"=ROUND(AVERAGE(I{row},N{row}),2)"
    ws[f"L{row}"] = f"=ROUND(J{row}/2,0)"
    ws[f"M{row}"] = f"=ROUND(H{row}/G{row},2)"
    ws[f"N{row}"] = f"=ROUND(M{row}*1.5%+M{row},2)"
    ws[f"O{row}"] = f"=(N{row}-B{row})/B{row}"
    ws[f"P{row}"] = f"=(I{row}-B{row})*J{row}"
    ws[f"Q{row}"] = f"=ROUND((K{row}-I{row})*L{row},2)"
    ws[f"R{row}"] = f"=ROUND(S{row}-H{row},2)"
    ws[f"S{row}"] = f"=ROUND(H{row}*1.5%+H{row},2)"
    ws[f"T{row}"] = f"=ROUND((H{row}*D{row}+H{row})-H{row},2)"

# Save to Excel
wb.save("output.xlsx")
print(f"Excel file 'output.xlsx' generated with U1 = {u1_value}. DOWN starts from D3 at 0.25%, all values rounded to 2 decimals.")



# def on_data(wsapp, message):
            #     # --- Real Market Mode (Uncomment when market is open) ---
            #     # token = message.get('token')
            #     # ltp = message.get('last_traded_price', 0) / 100
                
            #     # --- Simulation Mode (Comment out when market is open) ---
            #     token = message.get('token')
            #     with app.app_context():
            #         stock = Stock.query.filter_by(symboltoken=token).first()
            #         if stock:
            #             last_trade = Trade.query.filter_by(stock_symbol=stock.tradingsymbol, user_email=user.email).order_by(Trade.id.desc()).first()
            #             base_price = last_trade.entry_price if last_trade else 100.0
            #             ltp = base_price * (1 + random.uniform(-0.05, 0.05))
            #             message['last_traded_price'] = ltp * 100
            #             # --- End of Simulation Mode ---

            #             message['name'] = stock.tradingsymbol
            #             live_prices[token] = {'price': ltp, 'name': stock.tradingsymbol}
            #             try:
            #                 logger.debug(f"Processing strategy for {stock.tradingsymbol} with LTP={ltp}")
            #                 process_strategy(user, stock.tradingsymbol, ltp, smart_api)
            #             except Exception as e:
            #                 logger.error(f"Error in process_strategy for {stock.tradingsymbol}: {str(e)}")
            #                 raise
            #             socketio.emit('stock_stream', {'message': 'New tick (Simulated)' if 'base_price' in locals() else 'New tick','data': message}, namespace='/stream', to=user_email)