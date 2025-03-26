import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.comments import Comment
import os

def analyze_excel_sheet(file_path, output_analysis_file="excel_analysis.txt"):
    try:
        # Load the workbook and get the active sheet
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Get sheet dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Convert to pandas DataFrame
        df = pd.read_excel(file_path)
        
        # Open output file for writing
        with open(output_analysis_file, 'w', encoding='utf-8') as f:
            # Write basic sheet information
            f.write("Sheet Analysis Report\n")
            f.write("=====================\n")
            f.write(f"Number of rows: {max_row}\n")
            f.write(f"Number of columns: {max_col}\n")
            f.write(f"Column headers: {list(df.columns)}\n")
            f.write("\nData Preview:\n")
            f.write(str(df.head()) + "\n")
            
            # Dictionary to store formula information
            formula_cells = {}
            
            # Analyze each cell for formulas
            f.write("\nFormula Detection:\n")
            f.write("==================\n")
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell_address = cell.coordinate
                    
                    # Check if cell contains a formula
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_cells[cell_address] = cell.value
                        f.write(f"Cell {cell_address}: Formula found - {cell.value}\n")
                        
                        # Add a comment to the cell
                        comment = Comment(
                            f"Formula detected: {cell.value}\nAnalyzed on: {pd.Timestamp.now()}",
                            "Grok Analysis"
                        )
                        cell.comment = comment
            
            # Write data types for each column
            f.write("\nColumn Data Types:\n")
            f.write("==================\n")
            for column, dtype in df.dtypes.items():
                f.write(f"{column}: {dtype}\n")
            
            # Write basic statistics
            f.write("\nBasic Statistics:\n")
            f.write("=================\n")
            f.write(str(df.describe()) + "\n")
            
            # Write all cell values
            f.write("\nAll Cell Values:\n")
            f.write("================\n")
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(cell_value)
                f.write(f"Row {row}: {row_data}\n")
            
            # Write summary
            f.write("\nSummary of Results:\n")
            f.write(f"Total formulas found: {len(formula_cells)}\n")
            f.write(f"Formula locations: {list(formula_cells.keys())}\n")
        
        # Save the workbook with added comments
        wb.save(file_path)
        print(f"Analysis complete!")
        print(f"Results saved to: {output_analysis_file}")
        print(f"Excel file with comments saved at: {file_path}")
        
        return {
            'row_count': max_row,
            'col_count': max_col,
            'formulas': formula_cells,
            'headers': list(df.columns),
            'data': df.to_dict()
        }
        
    except Exception as e:
        error_msg = f"An error occurred: {str(e)}"
        with open(output_analysis_file, 'w', encoding='utf-8') as f:
            f.write(error_msg)
        print(error_msg)
        return None

# Example usage
if __name__ == "__main__":
    # Specify your Excel file path
    excel_file = "strategy.xlsx"
    
    # Check if file exists
    if not os.path.exists(excel_file):
        print("Please provide a valid Excel file path")
    else:
        # Analyze the sheet and save results
        result = analyze_excel_sheet(
            excel_file,
            "excel_analysis_results.txt"  # Custom output file name
        )